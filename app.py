import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment

# 设置页面配置
st.set_page_config(page_title="综合成绩明细处理工具", layout="wide")
st.title("📊 综合成绩明细处理工具")

# 上传文件
uploaded_file = st.file_uploader("请上传 '综合得分明细.xlsx' 文件", type=['xlsx'])

if uploaded_file is not None:
    try:
        # 读取Excel文件（所有列都作为字符串读取，空值保持为空）
        df_raw = pd.read_excel(uploaded_file, header=None, dtype=str, keep_default_na=False)
        
        if len(df_raw) < 3:
            st.error("文件格式似乎不正确，请确保文件至少包含标题、字段名和一行数据。")
            st.stop()

        # 获取原始文件的标题
        original_title = str(df_raw.iloc[0, 0]) if len(df_raw) > 0 else "综合得分明细"

        # 以第二行（索引1）作为字段名
        df = df_raw.copy()
        raw_columns = df.iloc[1].tolist()
        clean_columns = []
        for col in raw_columns:
            if pd.isna(col) or str(col).lower() == 'nan':
                clean_columns.append('')
            else:
                col_str = str(col)
                if col_str.endswith('.0') and col_str[:-2].isdigit():
                    clean_columns.append(col_str[:-2])
                else:
                    clean_columns.append(col_str)

        df.columns = clean_columns
        df = df.iloc[2:].reset_index(drop=True)  # 从第三行（索引2）开始是数据

        # 清洗数据
        df = df.replace(['nan', 'NaN', 'NAN', 'None'], '')
        df = df.fillna('')

        # 提取字段
        fixed_fields = ['地区', '学校', '班级', '姓名', '考号']
        existing_fixed_fields = [f for f in fixed_fields if f in df.columns]
        
        all_columns = df.columns.tolist()
        dynamic_fields = [col for col in all_columns if col not in fixed_fields and col != '总分']

        st.success(f"文件读取成功！共读取到 {len(df)} 条数据。")

        # 使用表单包裹所有的输入和提交按钮
        with st.form("process_form"):
            st.header("1. 字段名修改")
            st.info("提示：输入新字段名则替换，不填则保持原字段名不变。")
            
            # 使用列布局显示重命名输入框
            rename_inputs = {}
            cols = st.columns(len(existing_fixed_fields) if existing_fixed_fields else 1)
            for i, field in enumerate(existing_fixed_fields):
                with cols[i]:
                    rename_inputs[field] = st.text_input(f"【{field}】修改为:", placeholder="回车保持原名")

            st.header("2. 题目分类标记")
            st.info("输入格式：直接输入序号（1/2/3），连续多个可用减号（如2-10），多个序号用英文逗号分隔（如 1,3,5）。未填写的题型将不会出现在最终表格中。")
            
            # 展示题目列表
            st.write(f"**共有 {len(dynamic_fields)} 道题目需要分类：**")
            questions_display = "  |  ".join([f"**{i+1}**.{f}" for i, f in enumerate(dynamic_fields)])
            st.markdown(f"> {questions_display}")
            
            # 题型输入框
            col1, col2 = st.columns(2)
            with col1:
                single_input = st.text_input("单选题序号", placeholder="例如: 1,3,5 或 1-10")
                judge_input = st.text_input("判断题序号", placeholder="例如: 21,22,23")
            with col2:
                multi_input = st.text_input("多选题序号", placeholder="例如: 11,13,15")
                subj_input = st.text_input("主观题序号", placeholder="例如: 30 或 主观题")

            # 提交按钮
            submitted = st.form_submit_button("开始处理生成报表", type="primary")

        if submitted:
            # --- 1. 处理重命名 ---
            rename_map = {k: v.strip() for k, v in rename_inputs.items() if v.strip()}
            if rename_map:
                df.rename(columns=rename_map, inplace=True)
                final_fixed_fields = [rename_map.get(f, f) for f in existing_fixed_fields]
            else:
                final_fixed_fields = existing_fixed_fields

            # --- 2. 处理题目分类解析 ---
            def parse_input(input_str, assigned):
                indices = set()
                if not input_str.strip():
                    return []
                
                for part in input_str.split(','):
                    part = part.strip()
                    if not part: continue
                    if '-' in part:
                        try:
                            start, end = map(int, part.split('-'))
                            indices.update(range(start, end + 1))
                        except:
                            st.warning(f"无法解析范围 '{part}'，已跳过")
                    else:
                        try:
                            indices.add(int(part))
                        except:
                            # 尝试按字段名匹配
                            matched = False
                            for idx, field in enumerate(dynamic_fields, 1):
                                if field == part:
                                    indices.add(idx)
                                    matched = True
                                    break
                            if not matched:
                                st.warning(f"无法解析 '{part}'，已跳过")
                
                result = []
                for idx in sorted(indices):
                    if idx < 1 or idx > len(dynamic_fields):
                        st.warning(f"序号 {idx} 超出范围，已跳过")
                        continue
                    field_name = dynamic_fields[idx - 1]
                    if field_name in assigned:
                        st.warning(f"'{field_name}' 已被其他题型占用，已跳过")
                        continue
                    result.append(field_name)
                    assigned.add(field_name)
                return result

            assigned_fields = set()
            category_map = {
                '单选': parse_input(single_input, assigned_fields),
                '多选': parse_input(multi_input, assigned_fields),
                '判断': parse_input(judge_input, assigned_fields),
                '主观题': parse_input(subj_input, assigned_fields)
            }

            # 动态识别哪些题型有数据
            active_obj_types = [cat for cat in ['单选', '多选', '判断'] if len(category_map[cat]) > 0]
            has_subj = len(category_map['主观题']) > 0

            with st.spinner('正在生成表格，请稍候...'):
                
                def calc_score(cat_list, row_data_source):
                    score = 0
                    for col in cat_list:
                        val = row_data_source.get(col, '')
                        if val and str(val).replace('.', '').replace('-', '').isdigit():
                            score += float(val)
                    return score

                # ========== 1. 客观明细表数据 ==========
                objective_columns = final_fixed_fields + active_obj_types + (['客观分'] if active_obj_types else [])
                objective_data = []
                for idx, row in df.iterrows():
                    row_data = {f: row.get(f, '') for f in final_fixed_fields}
                    obj_total = 0
                    for cat in active_obj_types:
                        cat_score = calc_score(category_map[cat], row)
                        row_data[cat] = cat_score
                        obj_total += cat_score
                    
                    if active_obj_types:
                        row_data['客观分'] = obj_total
                    
                    objective_data.append(row_data)
                df_objective = pd.DataFrame(objective_data, columns=objective_columns)

                # ========== 2. 主观明细表数据 ==========
                subjective_columns = final_fixed_fields + category_map['主观题'] + (['主观分'] if has_subj else [])
                subjective_data = []
                for idx, row in df.iterrows():
                    row_data = {f: row.get(f, '') for f in final_fixed_fields}
                    if has_subj:
                        subj_score = 0
                        for col in category_map['主观题']:
                            val = row.get(col, '')
                            row_data[col] = val
                            if val and str(val).replace('.', '').replace('-', '').isdigit():
                                subj_score += float(val)
                        row_data['主观分'] = subj_score
                    subjective_data.append(row_data)
                df_subjective = pd.DataFrame(subjective_data, columns=subjective_columns)

                # ========== 3. 主客观明细表数据 ==========
                main_columns = final_fixed_fields + active_obj_types + category_map['主观题'] + ['总分']
                main_data = []
                for i in range(len(df)):
                    row_data = {f: objective_data[i][f] for f in final_fixed_fields}
                    
                    for cat in active_obj_types:
                        row_data[cat] = objective_data[i][cat]
                    
                    for col in category_map['主观题']:
                        row_data[col] = subjective_data[i][col]
                    
                    # 计算总分（安全获取客观分和主观分，防止没有这些题型时报错）
                    obj_score = objective_data[i].get('客观分', 0)
                    subj_score = subjective_data[i].get('主观分', 0)
                    row_data['总分'] = obj_score + subj_score
                    
                    main_data.append(row_data)
                df_main = pd.DataFrame(main_data, columns=main_columns)

                # ========== 4. 主客观简表数据 ==========
                simple_columns = final_fixed_fields.copy()
                if active_obj_types: simple_columns.append('客观分')
                if has_subj: simple_columns.append('主观分')
                simple_columns.append('总分')
                
                simple_data = []
                for i in range(len(df)):
                    row_data = {f: objective_data[i][f] for f in final_fixed_fields}
                    if active_obj_types: 
                        row_data['客观分'] = objective_data[i]['客观分']
                    if has_subj: 
                        row_data['主观分'] = subjective_data[i]['主观分']
                        
                    obj_score = objective_data[i].get('客观分', 0)
                    subj_score = subjective_data[i].get('主观分', 0)
                    row_data['总分'] = obj_score + subj_score
                    
                    simple_data.append(row_data)
                df_simple = pd.DataFrame(simple_data, columns=simple_columns)

                # ========== 5. 总分表数据 ==========
                total_columns = ['序号'] + final_fixed_fields + ['总分', '备注']
                total_data = []
                for i in range(len(df)):
                    row_data = {'序号': i + 1}
                    row_data.update({f: objective_data[i][f] for f in final_fixed_fields})
                    
                    obj_score = objective_data[i].get('客观分', 0)
                    subj_score = subjective_data[i].get('主观分', 0)
                    row_data['总分'] = obj_score + subj_score
                    row_data['备注'] = ''
                    
                    total_data.append(row_data)
                df_total = pd.DataFrame(total_data, columns=total_columns)

                # ========== 使用openpyxl写入并生成文件流 ==========
                output = BytesIO()
                wb = Workbook()

                def write_sheet(ws, title, columns, df_data):
                    if not columns: # 防止因为没有任何列导致合并单元格报错
                        return
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
                    title_cell = ws.cell(row=1, column=1, value=title)
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    for col_idx, col_name in enumerate(columns, 1):
                        ws.cell(row=2, column=col_idx, value=col_name)
                    
                    for row_idx, row_vals in enumerate(df_data.values, 3):
                        for col_idx, value in enumerate(row_vals, 1):
                            if isinstance(value, (int, float)):
                                ws.cell(row=row_idx, column=col_idx, value=value)
                            else:
                                val_str = str(value) if value is not None else ''
                                if val_str.lower() in ['nan', 'none', 'nat']:
                                    val_str = ''
                                ws.cell(row=row_idx, column=col_idx, value=val_str)

                # 写入5个表
                ws1 = wb.active
                ws1.title = "客观明细表"
                write_sheet(ws1, original_title, objective_columns, df_objective)

                ws2 = wb.create_sheet(title="主观明细表")
                write_sheet(ws2, original_title, subjective_columns, df_subjective)

                ws3 = wb.create_sheet(title="主客观明细表")
                write_sheet(ws3, original_title, main_columns, df_main)

                ws4 = wb.create_sheet(title="主客观简表")
                write_sheet(ws4, original_title, simple_columns, df_simple)

                ws5 = wb.create_sheet(title="总分表")
                write_sheet(ws5, original_title, total_columns, df_total)

                wb.save(output)
                processed_data = output.getvalue()

            st.success("🎉 处理完成！未填写的题型已自动在表格中隐藏。请点击下方按钮下载。")
            
            st.download_button(
                label="📥 下载综合成绩表.xlsx",
                data=processed_data,
                file_name="综合成绩表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

    except Exception as e:
        st.error(f"处理文件时发生错误: {e}")
